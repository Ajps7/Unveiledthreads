"""Persistence and spreadsheet export.

The CSV column order is fixed by the brief and asserted in the tests, because
downstream sheets and imports break silently when columns move. Two columns are
appended beyond the brief's list: `Outreach Note` (required by the brief itself,
just not listed in the column spec) and `Last Verified`.
"""

from __future__ import annotations

import csv
import json
import os
from datetime import date
from typing import Iterable, Optional

from .models import Lead, NOT_FOUND, dump_leads, load_leads

COLUMNS: list[str] = [
    "Brand Name", "Website", "Instagram Handle", "Instagram URL", "Contact Email",
    "Founder/Owner", "Founder Contact/Profile", "UK Location", "Category",
    "Instagram Followers", "Ecommerce Website", "Existing Marketplaces",
    "Lead Score", "Priority", "Contact Page", "Source URLs", "Notes",
    "Outreach Note", "Last Verified",
]

_FIELD_BY_COLUMN = {
    "Brand Name": "brand_name",
    "Website": "website",
    "Instagram Handle": "instagram_handle",
    "Instagram URL": "instagram_url",
    "Contact Email": "contact_email",
    "Founder/Owner": "founder_owner",
    "Founder Contact/Profile": "founder_profile",
    "UK Location": "uk_location",
    "Category": "category",
    "Instagram Followers": "instagram_followers",
    "Ecommerce Website": "sells_own_website",
    "Existing Marketplaces": "existing_marketplaces",
    "Contact Page": "contact_page",
}


def lead_to_row(lead: Lead) -> dict[str, str]:
    row = {col: getattr(lead, attr, NOT_FOUND) or NOT_FOUND
           for col, attr in _FIELD_BY_COLUMN.items()}
    row["Lead Score"] = str(lead.lead_score) if lead.lead_score is not None else NOT_FOUND
    row["Priority"] = lead.priority or NOT_FOUND
    row["Source URLs"] = " | ".join(lead.all_source_urls()) or NOT_FOUND
    row["Notes"] = lead.notes or ""
    row["Outreach Note"] = lead.outreach_note or ""
    row["Last Verified"] = lead.last_verified or NOT_FOUND
    return row


def export_csv(leads: Iterable[Lead], path: str) -> str:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS)
        writer.writeheader()
        for lead in leads:
            writer.writerow(lead_to_row(lead))
    return path


def export_xlsx(leads: Iterable[Lead], path: str) -> Optional[str]:
    """Write an .xlsx if openpyxl is installed; return None if it is not."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    leads = list(leads)
    wb = Workbook()
    ws = wb.active
    ws.title = "UK Independent Brands"
    ws.append(COLUMNS)

    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    priority_fill = {
        "HIGH PRIORITY": PatternFill("solid", fgColor="D1FAE5"),
        "MEDIUM PRIORITY": PatternFill("solid", fgColor="FEF3C7"),
        "LOW PRIORITY": PatternFill("solid", fgColor="FEE2E2"),
    }
    priority_col = COLUMNS.index("Priority") + 1

    for lead in leads:
        row = lead_to_row(lead)
        ws.append([row[c] for c in COLUMNS])
        fill = priority_fill.get(row["Priority"])
        if fill:
            ws.cell(row=ws.max_row, column=priority_col).fill = fill

    widths = {"Brand Name": 26, "Website": 34, "Instagram Handle": 22,
              "Instagram URL": 38, "Contact Email": 30, "Founder/Owner": 24,
              "Founder Contact/Profile": 38, "UK Location": 20, "Category": 30,
              "Source URLs": 60, "Notes": 70, "Outreach Note": 80}
    for idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 16)

    wb.save(path)
    return path


def export_markdown(leads: Iterable[Lead], path: str) -> str:
    """Human-readable review sheet — easier to sanity-check than a CSV."""
    leads = list(leads)
    lines = [
        "# UnveiledThreads — UK Independent Brand Prospects",
        "",
        f"Generated {date.today().isoformat()} · {len(leads)} leads",
        "",
        "| # | Brand | Location | Score | Priority | IG | Email |",
        "|---|-------|----------|-------|----------|----|-------|",
    ]
    for i, lead in enumerate(leads, start=1):
        lines.append(
            f"| {i} | {lead.brand_name} | {lead.uk_location} | {lead.lead_score} | "
            f"{lead.priority} | {lead.instagram_handle} | {lead.contact_email} |"
        )
    lines.append("")
    for lead in leads:
        lines += [
            f"## {lead.brand_name} — {lead.lead_score}/100 ({lead.priority})",
            "",
            f"- **Website:** {lead.website}",
            f"- **Instagram:** {lead.instagram_handle} ({lead.instagram_url}) — "
            f"{lead.instagram_followers} followers",
            f"- **Contact:** {lead.contact_email} · {lead.contact_page}",
            f"- **Founder:** {lead.founder_owner} · {lead.founder_profile}",
            f"- **Location:** {lead.uk_location}",
            f"- **Category:** {lead.category}",
            f"- **Own-site ecommerce:** {lead.sells_own_website} · "
            f"**Marketplaces:** {lead.existing_marketplaces}",
            f"- **Outreach:** {lead.outreach_note}",
            f"- **Notes:** {lead.notes}",
            f"- **Sources:** {', '.join(lead.all_source_urls()) or NOT_FOUND}",
            "",
        ]
    content = "\n".join(lines)
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(content)
    return path


class LeadStore:
    """JSON-backed store with idempotent upsert, so repeat runs accumulate."""

    def __init__(self, path: str):
        self.path = path

    def load(self) -> list[Lead]:
        if not os.path.exists(self.path):
            return []
        return load_leads(self.path)

    def save(self, leads: list[Lead]) -> None:
        os.makedirs(os.path.dirname(os.path.abspath(self.path)), exist_ok=True)
        dump_leads(leads, self.path)

    def upsert(self, incoming: list[Lead]) -> list[Lead]:
        """Merge new findings into the stored set without losing verified data."""
        from .dedup import deduplicate

        combined = self.load() + incoming
        merged = deduplicate(combined)
        self.save(merged)
        return merged

    def stats(self) -> dict:
        leads = self.load()
        from .scoring import summarise
        payload = summarise(leads)
        payload["with_email"] = sum(1 for l in leads if l.contact_email != NOT_FOUND)
        payload["with_founder"] = sum(1 for l in leads if l.founder_owner != NOT_FOUND)
        return payload
